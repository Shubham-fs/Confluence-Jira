"""Excel export of the reports using openpyxl."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F6FEB")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for row in rows:
        ws.append(row)
    # Auto-ish width based on content length.
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in rows:
            value = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)


def build_assigned_workbook(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Assigned Issues"
    headers = ["Key", "Summary", "Status", "Assignee", "Reporter", "Created", "Updated"]
    rows = [
        [
            i.get("key"),
            i.get("summary"),
            i.get("status"),
            i.get("assignee"),
            i.get("reporter"),
            i.get("created"),
            i.get("updated"),
        ]
        for i in report.get("issues", [])
    ]
    _write_sheet(ws, headers, rows)
    return _to_bytes(wb)


def build_transitions_workbook(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Forward Transitions"
    headers = [
        "Key",
        "Summary",
        "From",
        "To",
        "Transitioned At",
        "Performed By",
        "Assignee",
    ]
    rows = [
        [
            i.get("key"),
            i.get("summary"),
            i.get("from_status"),
            i.get("to_status"),
            i.get("transitioned_at"),
            i.get("performed_by"),
            i.get("assignee"),
        ]
        for i in report.get("issues", [])
    ]
    _write_sheet(ws, headers, rows)
    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
